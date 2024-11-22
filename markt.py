#=================[Importing Libraries]===========================
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import cv2
from pyzbar.pyzbar import decode
from PIL import Image, ImageTk
import openpyxl as xl
import pyttsx3
from tkinter.messagebox import*
import datetime
 

#===========[Creating the Main Window]===============

window=Tk()
window.geometry("1150x690+100+0")
window.config(bg="black")
window.title("super market")
window.iconbitmap("img/ico/market.ico")

# Global Variables
products =[]#pro1
product_prices=[]#pro2
temporary_prices=[]#pro

current_time = datetime.datetime.now()
current_date  = current_time.strftime("%Y-%m-%d")

#=================[Helper Functions]===========================
# Close Camera Function
def close_camera():
     global is_camera_open
     is_camera_open=True     

# close window
def close_window():
     window.destroy()   
def clear_fields_and_table():
     entry_buyer_name.delete("0",END)#e
     entry_buyer_number.delete("0",END)#e1
     entry_buyer_address.delete("0",END)#e2
     entry_total_amount.delete("0",END)#e3
     entry_purchase_date.delete("0",END)#e4
     barcode_entry.delete("0",END)#lc2
     product_name_entry.delete("0",END)#lc4
     product_price_entry.delete("0",END)#lc6
     for item in tree_view.get_children():
        tree_view.delete(item)#trv
def clear_fields():
     entry_buyer_name.delete("0",END)
     entry_buyer_number.delete("0",END)
     entry_buyer_address.delete("0",END)
     entry_total_amount.delete("0",END)
     entry_purchase_date.delete("0",END)
def search_buyer():
     buyer_name =entry_buyer_name.get()
     workbook =xl.load_workbook("customers.xlsx")
     sheet =workbook ['h']
     found = False

     for j in range(2,sheet.max_row+1):
               call=sheet.cell(j,1)
               c=str(a)
               
               if c==str(call.value):
                found = True
                entry_buyer_name.delete("0",END)
                entry_buyer_number.delete("0",END)
                entry_buyer_address.delete("0",END)
                entry_total_amount.delete("0",END)
                entry_purchase_date.delete("0",END)

                entry_buyer_name=str(sheet.cell(j,1).value)#dd
                entry_buyer_number=str(sheet.cell(j,2).value)#dd1
                entry_buyer_address=str(sheet.cell(j,3).value)#dd2
                entry_total_amount=str(sheet.cell(j,4).value)#dd3
                entry_purchase_date=str(sheet.cell(j,5).value)#dd4

                entry_buyer_name.insert(0,entry_buyer_name)
                entry_buyer_number.insert(0,entry_buyer_number)
                entry_buyer_address.insert(0,entry_buyer_address)
                entry_total_amount.insert(0,entry_total_amount)
                entry_purchase_date.insert(0,entry_purchase_date)
               
               
     if found == False:
         showerror(title='خطأ ',message='أسم المشتري غير موجود')
     
def delete_invoice():
     
     buyer_name = entry_buyer_name.get()
     workbook = xl.load_workbook("customers.xlsx")
     sheet = workbook['h']
     found = False
     for j in range(2,sheet.max_row+1):
               
               buyer_cell =sheet.cell(j,1)
               c=str(a)
               
               if c==str(buyer_cell.value):
                found= True
                entry_buyer_name.delete("0",END)
                entry_buyer_number.delete("0",END)
                entry_buyer_address.delete("0",END)
                entry_total_amount.delete("0",END)
                entry_purchase_date.delete("0",END)

                dd=sheet.cell(j,1)
                dd1=sheet.cell(j,2)
                dd2=sheet.cell(j,3)
                dd3=sheet.cell(j,4)
                dd4=sheet.cell(j,5)

                entry_buyer_name.insert(0,dd)
                entry_buyer_number.insert(0,dd1)
                entry_buyer_address.insert(0,dd2)
                entry_total_amount.insert(0,dd3)
                entry_purchase_date.insert(0,dd4)
                showwarning(title='تحذير ',message='هل أنت متأكد ')
     
                dd.value=""
                dd1.value=""
                dd2.value=""
                dd3.value=""
                dd4.value=""
                entry_buyer_name.delete("0",END)
                entry_buyer_number.delete("0",END)
                entry_buyer_address.delete("0",END)
                entry_total_amount.delete("0",END)
                entry_purchase_date.delete("0",END)
                showinfo(title='تم الحذف',message='تم حذف الفاتورة بناجاح')
                workbook.save("customers.xlsx")
     if found==True:
         showerror(title='خطأ ',message='أسم المشتري غير موجود')
                 
def h():
     s=barcode_entry.get()
     s1=product_name_entry.get()
     s2=product_price_entry.get()
     wb=xl.load_workbook("excel/ib.xlsx")
     sheet=wb['ibe']
     c=(sheet.max_row +1)
     cell=sheet.cell(c,1)
     cell1=sheet.cell(c,2)
     cell2=sheet.cell(c,3)
     cell.value=s
     cell1.value=s1
     cell2.value=int(s2)
     wb.save("excel/ib.xlsx")
     showinfo(title='تم الحفظ',message='تم حفظ المنتج بناجاح')
def v():
     tree_view.insert(parent="",index=0,values=(m,n,str(code)))
def itemm(ferst,scnd):
     tree_view.insert(parent="",index=0,values=(ferst,scnd))

   
def cam():
    global f
    global m
    global n
    global code
    
    
    
    
    f=False
    # Open the camera
    cap = cv2.VideoCapture(0)
    c=True
    while c:
        ret, frame = cap.read()
        # Convert the frame to grayscale for barcode detection
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        if f==True:
           del(cap)
           f=False

        # Decode barcodes in the frame
        decoded_objects = decode(frame)

        for obj in decoded_objects:
            pyttsx3.speak('tit')
            
            barcode_entry.delete('0', END)
            product_name_entry.delete('0', END)
            product_price_entry.delete('0', END)


            
            code = str(obj.data.decode('utf-8'))
            barcode_entry.insert(0,code)
            
            #l.config(text= code)
            wb=xl.load_workbook("excel/ib.xlsx")
            sheet=wb['ibe']
            for d in range(2,sheet.max_row+1):
               call=sheet.cell(d,1)
               
               if code==str(call.value):
                m=str(sheet.cell(d,2).value)
                n=str(sheet.cell(d,3).value)
                product_name_entry.insert(0,m)
                #lc2.insert("1",code)
                product_price_entry.insert(0,n)
                temporary_prices.append(n)
                
                itemm(m,n)
                
                          
                            
                
                
                
                            
               
              
        # Display the camera feed in the Tkinter window
        cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGBA)
        img = Image.fromarray(cv2image)
        img_tk = ImageTk.PhotoImage(image=img)
        anvas.create_image(0, 0, image=img_tk, anchor=NW)
        window.update()
        
           
           

    # Release the camera
    cap.release()
def cc():
     global p

     p=0
     entry_buyer_number.delete("0",END)
     entry_total_amount.delete("0",END)
     entry_purchase_date.delete("0",END)

     for j in temporary_prices:
         k=int(j)
         p=(p+k)
     
     p2=0
     for j2 in product_prices:
         k2=float(j2)
         p2=(p2+k2)
    
     p3=0
     for j3 in products:
         k3=float(j3)
         p3=(p3+k3)
     d1=p+p2+p3 
     wn=xl.load_workbook("excel/a.xlsx")
     sheet1=wn['n'] 
     call1=sheet1.cell(1,1)
     c=str(call1.value)

     entry_buyer_number.insert("1",c)
     entry_total_amount.insert("1",d1)
     entry_purchase_date.insert('1',str(current_date))
      
def vv():
     im=entry_buyer_name.get()
     im1=entry_buyer_number.get()
     im2=entry_buyer_address.get()
     im3=entry_total_amount.get()
     im4=entry_purchase_date.get()
     ws=xl.load_workbook("customers.xlsx")
     sheet1=ws['h']
     nn=(sheet1.max_row +1)
     call1=sheet1.cell(nn,1)
     call2=sheet1.cell(nn,2)
     call3=sheet1.cell(nn,3)
     call4=sheet1.cell(nn,4)
     call5=sheet1.cell(nn,5)

     call1.value=im
     call2.value=im1
     call3.value=im2
     call4.value=im3
     call5.value=im4

     showinfo(title='تم الحفظ',message='تم حفظ الفاتورة بناجاح')
     wn=xl.load_workbook("excel/a.xlsx")
     sheet1=wn['n'] 
     n=sheet1.cell(1,1)
     c=n.value
     n.value=c=c+1
     wn.save("excel/a.xlsx")
     ws.save("customers.xlsx")
def xx(e):
    for item in trvv.get_children():
        trvv.delete(item)
    dd=ee.get()
    wb=xl.load_workbook("excel/ib.xlsx")
    sheet=wb['ibe']
    for d in range(2,sheet.max_row+1):
        call=sheet.cell(d,1)
        m=str(sheet.cell(d,2).value)
        n=str(sheet.cell(d,3).value)
               
        z=m.find(dd)
        if z != -1:
            trvv.insert(parent="",index=0,values=(m,n))
                
 
def git_slct(e):
    itt=trvv.focus() 
    ss=trvv.item(itt) 
    it1=ss.get("values")[0]
    it2=ss.get("values")[1]
    itemm(it1,it2)
    products.append(it2)
def xxx():
   global ee
   global trvv
   w2=Tk()
   w2.geometry("400x600+0+0")
   w2.iconbitmap("img/ico/market.ico")
   w2.title("بحث عن منتج")

   ee=Entry(w2,width=100,bg="grey",fg="white")
   ee.pack()
   ee.bind('<Return>',xx)

   trvv=ttk.Treeview(w2,height=600,columns=("1","2"),show="headings")
   trvv.heading("1", text ="المواد")
   trvv.heading("2", text ="السعر")
  #trv.heading("2", text ="العدد",anchor='c')
  #trv.heading("2", text ="الحساب الكلي",anchor='c')
   trvv.pack()
   trvv.bind('<<TreeviewSelect>>',git_slct)
def ccc1():
    
    enn1=en1.get()
    enn2=int(en2.get())
    enn3=int(en3.get())
    enn4=en4.get()
    
    en4.delete('0', END)
    gg=enn2/1000
    gg1=gg*enn3

    en4.insert(0,gg1)

def ccc2():
    en1.delete('0', END)
    en2.delete('0', END)
    en3.delete('0', END)
    en4.delete('0', END)
def ccc3():
    enn1=en1.get()
    enn2=en2.get()
    enn3=en3.get()
    enn4=en4.get()

    ww=str(enn4)

    tree_view.insert(parent="",index=0,values=(enn1,enn4))
    product_prices.append(ww)
       
def ccc():
    global en1
    global en2
    global en3
    global en4

    w3=Tk()
    w3.iconbitmap("img/ico/market.ico")
    w3.geometry("500x400")
    w3.title(" وزن البقوليات ")
    w3.config(bg="grey")
    
    la1=Label(w3,bg="blue",width=47,text="أدخل أسم المنتج",fg="white")
    la1.place(x=80,y=5)

    en1=Entry(w3,width=55,justify="center")
    en1.place(x=80,y=30)

    la2=Label(w3,bg="blue",width=47,text="أدخل سعر الكيلو",fg="white")
    la2.place(x=80,y=55)

    en2=Entry(w3,width=55,justify="center")
    en2.place(x=80,y=80)

    la3=Label(w3,bg="blue",width=47,text="أدخل وزن المنتج بلغرام",fg="white")
    la3.place(x=80,y=105)
 
    en3=Entry(w3,width=55,justify="center")
    en3.place(x=80,y=130)

    la4=Label(w3,bg="blue",width=47,text="سعر المنتج",fg="white")
    la4.place(x=80,y=155)
 
    en4=Entry(w3,width=55,justify="center")
    en4.place(x=80,y=175)

    ba1=Button(w3,text="حساب سعر المنتج",bg="green",fg="white",height=3,command=ccc1)
    ba1.place(x=120,y=200)

    ba2=Button(w3,text="أفراغ الحقول",bg="green",fg="white",height=3,command=ccc2)
    ba2.place(x=220,y=200)
    
    ba2=Button(w3,text="شراء المنتج",bg="green",fg="white",height=3,command=ccc3)
    ba2.place(x=290,y=200)







 
    
    
    
  #========================================================
a=PhotoImage(file="img/png/logo4.png")
f1=Frame(width=335,height=700,bg="grey") 
f1.place(x=812,y=0) 

f2=Frame(window,width=300,height=700,bg="grey") 
f2.place(x=405,y=70) 
  
f3=Frame(window,width=807,height=65,bg="blue") 
f3.place(x=0,y=0) 

f4=Frame(window,width=400,height=700,bg="grey") 
f4.place(x=0,y=70) 
  
f5=Frame(f4,width=260,height=395,bg="grey")
f5.place(x=75,y=390)

anvas = Canvas( f1,width=300, height=400,bg="white")
anvas.place(x=17,y=40) 
lc=Label(f1,text="Barcode Scanner",fg="white",bg="blue",font=("Stencil", "15", "bold italic"),width=30)
lc.place(x=0,y=0)


lc1=Label(f1,text="الرقم",width=40,fg="white",bg="blue")
lc1.place(x=27,y=460)
barcode_entry=Entry(f1,width=47,justify=CENTER)
barcode_entry.place(x=27,y=480)

lc3=Label(f1,text="المنتج",width=40,fg="white",bg="blue")
lc3.place(x=27,y=510)
product_name_entry=Entry(f1,width=47,justify=CENTER)
product_name_entry.place(x=27,y=530)

lc5=Label(f1,text="السعر",width=40,fg="white",bg="blue")
lc5.place(x=27,y=560)
product_price_entry=Entry(f1,width=47,justify=CENTER)
product_price_entry.place(x=27,y=580)

tree_view=ttk.Treeview(f2,height=700,columns=("1","2"),show="headings")
tree_view.heading("1", text ="المواد",anchor='center')
tree_view.heading("2", text ="السعر",anchor='center')



  
tree_view.pack()


  # Button to start scanning
l=Label(f4,text="أسم المشتري",bg="blue",fg="white",width=40)
l.place(x=60,y=10)
entry_buyer_name=Entry(f4,width=47,justify="center")
entry_buyer_name.place(x=60,y=30)

l1=Label(f4,text="رقم المشتري",bg="blue",fg="white",width=40)
l1.place(x=60,y=60)
entry_buyer_number=Entry(f4,width=47,justify="center")
entry_buyer_number.place(x=60,y=80)

l2=Label(f4,text="عنوان المشتري",bg="blue",fg="white",width=40)
l2.place(x=60,y=110)
entry_buyer_address=Entry(f4,width=47,justify="center")
entry_buyer_address.place(x=60,y=130)

l3=Label(f4,text="الحساب الكلي",bg="blue",fg="white",width=40)
l3.place(x=60,y=160)
entry_total_amount=Entry(f4,width=47,justify="center")
entry_total_amount.place(x=60,y=180)

l4=Label(f4,text="تريخ الشراء",bg="blue",fg="white",width=40)
l4.place(x=60,y=210)
entry_purchase_date=Entry(f4,width=47,justify="center")
entry_purchase_date.place(x=60,y=230)

s=Button(f4,text="حفظ الفاتورة",bg="blue",fg="white",width=40,command=vv)
s.place(x=60,y=270)

s2=Button(f4,text="أفراغ الحقول",bg="blue",fg="white",width=40,command=clear_fields)
s2.place(x=60,y=300)

s3=Button(f4,text="بحث عن المشتري",bg="blue",fg="white",width=40,command=search_buyer)
s3.place(x=60,y=330)

s=Button(f4,text="حذف  فاتورة",bg="blue",fg="white",width=40,command=delete_invoice)
s.place(x=60,y=360)




x=Button(f1,text="open camera",width=35,bg="blue",fg="white",command=cam)
x.place(x=40,y=608)
u=Button(f1,text="cloas camera",fg="white",bg="blue",width=35,command=close_camera)
u.place(x=40,y=635)

d=Button(f3,text="فاتورة جديدة",fg="white",bg="green",height=3,width=10,command=f)
d.place(x=720,y=6)
d2=Button(f3,text="حفظ المنتج",fg="white",bg="green",height=3,width=10,command=h)
d2.place(x=635,y=6)
o=Button(f3,text="شراء المنتجات ",fg="white",bg="green",height=3,width=10,command=cc)
o.place(x=550,y=6)
oo=Button(f3,text="أغلاق البرنامج  ",fg="white",bg="green",height=3,width=10,command=close_window)
oo.place(x=465,y=6)
ooo=Button(f3,text="بحث عن منتج",fg="white",bg="green",height=3,width=10,command=xxx)
ooo.place(x=380,y=6)
FF=Button(f3,text="منتج بلوزن",fg="white",bg="green",height=3,width=10,command=ccc)
FF.place(x=295,y=6)

bb=Button(f5,image=a)
bb.place(x=0,y=0)
  
  


window.mainloop()
  
